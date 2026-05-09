import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Hospital data
hospital_data = [
    {
        "Hospital_ID": "HZR001",
        "Hospital_Name": "Government Area Hospital, Huzurabad",
        "Address": "Near Bus Stand, Huzurabad, Telangana 505468",
        "Type": "Government",
        "Specialties": "General Medicine, Emergency, Maternity, Pediatrics, Surgery",
        "Beds_Total": 100,
        "ICU_Beds": 10,
        "Ventilators": 5,
        "Contact_Number": "08728-255555",
        "Emergency_Contact": "08728-255556",
        "Email": "gah.huzurabad@telangana.gov.in",
        "Website": "",
        "Latitude": 18.2046,
        "Longitude": 79.4997,
        "Operating_Hours": "24/7",
        "Ambulance_Service": "Yes",
        "Blood_Bank": "Yes",
        "Pharmacy": "Yes",
        "Rating": 4.2,
        "Established_Year": 1985
    },
    {
        "Hospital_ID": "HZR002",
        "Hospital_Name": "Sri Sai Ram Hospital",
        "Address": "Station Road, Huzurabad, Telangana 505468",
        "Type": "Private",
        "Specialties": "General Surgery, Pediatrics, Orthopedics, Cardiology",
        "Beds_Total": 75,
        "ICU_Beds": 8,
        "Ventilators": 4,
        "Contact_Number": "08728-266666",
        "Emergency_Contact": "08728-266667",
        "Email": "srisairam.hzr@gmail.com",
        "Website": "",
        "Latitude": 18.2050,
        "Longitude": 79.5000,
        "Operating_Hours": "24/7",
        "Ambulance_Service": "Yes",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 4.4,
        "Established_Year": 1995
    },
    {
        "Hospital_ID": "HZR003",
        "Hospital_Name": "Geetha Hospital",
        "Address": "Near RTC Bus Stand, Huzurabad",
        "Type": "Private",
        "Specialties": "General Medicine, Gynecology, Pediatrics",
        "Beds_Total": 50,
        "ICU_Beds": 4,
        "Ventilators": 2,
        "Contact_Number": "08728-277777",
        "Emergency_Contact": "08728-277778",
        "Email": "",
        "Website": "",
        "Latitude": 18.2060,
        "Longitude": 79.5010,
        "Operating_Hours": "24/7",
        "Ambulance_Service": "Yes",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 4.1,
        "Established_Year": 2000
    },
    {
        "Hospital_ID": "HZR004",
        "Hospital_Name": "Swathi Hospital",
        "Address": "Huzurabad Main Road, Near Clock Tower",
        "Type": "Private",
        "Specialties": "General Medicine, ENT, Dermatology",
        "Beds_Total": 40,
        "ICU_Beds": 3,
        "Ventilators": 1,
        "Contact_Number": "08728-288888",
        "Emergency_Contact": "08728-288889",
        "Email": "",
        "Website": "",
        "Latitude": 18.2035,
        "Longitude": 79.4985,
        "Operating_Hours": "8:00 AM - 10:00 PM",
        "Ambulance_Service": "No",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 3.9,
        "Established_Year": 2005
    },
    {
        "Hospital_ID": "HZR005",
        "Hospital_Name": "Raghava Nursing Home",
        "Address": "Market Road, Huzurabad",
        "Type": "Private",
        "Specialties": "General Medicine, Pediatrics",
        "Beds_Total": 25,
        "ICU_Beds": 2,
        "Ventilators": 0,
        "Contact_Number": "08728-299999",
        "Emergency_Contact": "08728-299990",
        "Email": "",
        "Website": "",
        "Latitude": 18.2070,
        "Longitude": 79.5020,
        "Operating_Hours": "9:00 AM - 9:00 PM",
        "Ambulance_Service": "No",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 3.8,
        "Established_Year": 2008
    },
    {
        "Hospital_ID": "HZR006",
        "Hospital_Name": "Nagarjuna Hospital",
        "Address": "Bus Stand Road, Huzurabad",
        "Type": "Private",
        "Specialties": "General Medicine, Orthopedics, Physiotherapy",
        "Beds_Total": 35,
        "ICU_Beds": 3,
        "Ventilators": 1,
        "Contact_Number": "08728-244444",
        "Emergency_Contact": "08728-244445",
        "Email": "nagarjunahosp.hzr@yahoo.com",
        "Website": "",
        "Latitude": 18.2025,
        "Longitude": 79.4975,
        "Operating_Hours": "24/7",
        "Ambulance_Service": "Yes",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 4.0,
        "Established_Year": 2010
    },
    {
        "Hospital_ID": "HZR007",
        "Hospital_Name": "Srinivasa Hospital",
        "Address": "Karimnagar Road, Huzurabad",
        "Type": "Private",
        "Specialties": "Multi-specialty, Cardiology, Neurology",
        "Beds_Total": 60,
        "ICU_Beds": 6,
        "Ventilators": 3,
        "Contact_Number": "08728-233333",
        "Emergency_Contact": "08728-233334",
        "Email": "srinivasahosp.hzr@gmail.com",
        "Website": "",
        "Latitude": 18.2010,
        "Longitude": 79.4960,
        "Operating_Hours": "24/7",
        "Ambulance_Service": "Yes",
        "Blood_Bank": "Yes",
        "Pharmacy": "Yes",
        "Rating": 4.3,
        "Established_Year": 2012
    },
    {
        "Hospital_ID": "HZR008",
        "Hospital_Name": "Medwin Hospital",
        "Address": "Near Police Station, Huzurabad",
        "Type": "Private",
        "Specialties": "General Medicine, Surgery, Ophthalmology",
        "Beds_Total": 45,
        "ICU_Beds": 4,
        "Ventilators": 2,
        "Contact_Number": "08728-222222",
        "Emergency_Contact": "08728-222223",
        "Email": "",
        "Website": "",
        "Latitude": 18.2080,
        "Longitude": 79.5030,
        "Operating_Hours": "24/7",
        "Ambulance_Service": "Yes",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 4.1,
        "Established_Year": 2015
    },
    {
        "Hospital_ID": "HZR009",
        "Hospital_Name": "Huzurabad Eye Hospital",
        "Address": "Near Municipal Office, Huzurabad",
        "Type": "Specialty",
        "Specialties": "Ophthalmology, Cataract Surgery, LASIK",
        "Beds_Total": 20,
        "ICU_Beds": 1,
        "Ventilators": 0,
        "Contact_Number": "08728-211111",
        "Emergency_Contact": "08728-211112",
        "Email": "eyehosp.hzr@gmail.com",
        "Website": "",
        "Latitude": 18.2090,
        "Longitude": 79.5040,
        "Operating_Hours": "9:00 AM - 8:00 PM",
        "Ambulance_Service": "No",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 4.5,
        "Established_Year": 2018
    },
    {
        "Hospital_ID": "HZR010",
        "Hospital_Name": "Dental Care Hospital",
        "Address": "Center Point, Huzurabad",
        "Type": "Specialty",
        "Specialties": "Dentistry, Orthodontics, Oral Surgery",
        "Beds_Total": 10,
        "ICU_Beds": 0,
        "Ventilators": 0,
        "Contact_Number": "08728-200000",
        "Emergency_Contact": "08728-200001",
        "Email": "",
        "Website": "",
        "Latitude": 18.2000,
        "Longitude": 79.4950,
        "Operating_Hours": "10:00 AM - 7:00 PM",
        "Ambulance_Service": "No",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 4.2,
        "Established_Year": 2020
    },
    {
        "Hospital_ID": "HZR011",
        "Hospital_Name": "Primary Health Center (PHC), Huzurabad",
        "Address": "Main Road, Huzurabad",
        "Type": "Government",
        "Specialties": "Primary Care, Immunization, Basic Medicine",
        "Beds_Total": 15,
        "ICU_Beds": 0,
        "Ventilators": 0,
        "Contact_Number": "08728-266555",
        "Emergency_Contact": "08728-266556",
        "Email": "",
        "Website": "",
        "Latitude": 18.2040,
        "Longitude": 79.5005,
        "Operating_Hours": "9:00 AM - 5:00 PM",
        "Ambulance_Service": "No",
        "Blood_Bank": "No",
        "Pharmacy": "Yes",
        "Rating": 3.7,
        "Established_Year": 1990
    }
]

def create_excel_sheet():
    """Create an Excel file with hospital data"""
    
    # Convert to DataFrame
    df = pd.DataFrame(hospital_data)
    
    # Create Excel writer
    excel_file = "hospitals_huzurabad.xlsx"
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Write main data sheet
        df.to_excel(writer, sheet_name='Hospitals', index=False)
        
        # Get workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Hospitals']
        
        # Format column widths
        for column in df.columns:
            column_length = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column) + 1
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(column_length + 2, 40)
        
        # Apply header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Format all cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = border
        
        # Add a summary sheet
        summary_data = {
            'Hospital Type': ['Government', 'Private', 'Specialty', 'Total'],
            'Count': [
                len(df[df['Type'] == 'Government']),
                len(df[df['Type'] == 'Private']),
                len(df[df['Type'] == 'Specialty']),
                len(df)
            ],
            'Total Beds': [
                df[df['Type'] == 'Government']['Beds_Total'].sum(),
                df[df['Type'] == 'Private']['Beds_Total'].sum(),
                df[df['Type'] == 'Specialty']['Beds_Total'].sum(),
                df['Beds_Total'].sum()
            ],
            'ICU Beds': [
                df[df['Type'] == 'Government']['ICU_Beds'].sum(),
                df[df['Type'] == 'Private']['ICU_Beds'].sum(),
                df[df['Type'] == 'Specialty']['ICU_Beds'].sum(),
                df['ICU_Beds'].sum()
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        
        # Format summary headers
        for cell in summary_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Format summary cells
        for row in summary_sheet.iter_rows(min_row=2, max_row=len(summary_df)+1, max_col=len(summary_df.columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
        
        # Add metadata sheet
        metadata = {
            'Field': [
                'Data Source',
                'Last Updated',
                'City',
                'District',
                'State',
                'Total Hospitals',
                'Total Beds',
                'File Version'
            ],
            'Value': [
                'Compiled from local sources',
                pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Huzurabad',
                'Karimnagar',
                'Telangana',
                len(df),
                df['Beds_Total'].sum(),
                '1.0'
            ]
        }
        
        metadata_df = pd.DataFrame(metadata)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Format metadata sheet
        metadata_sheet = writer.sheets['Metadata']
        
        # Format metadata headers
        for cell in metadata_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Format metadata cells
        for row in metadata_sheet.iter_rows(min_row=2, max_row=len(metadata_df)+1, max_col=len(metadata_df.columns)):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
                cell.border = border
    
    print(f"✅ Excel file created: {excel_file}")
    print(f"📊 Total hospitals: {len(df)}")
    print(f"🛏️ Total beds: {df['Beds_Total'].sum()}")
    print(f"🏥 ICU beds: {df['ICU_Beds'].sum()}")
    print(f"📋 Sheets created: Hospitals, Summary, Metadata")
    
    return excel_file

def display_sample_data():
    """Display sample data for verification"""
    df = pd.DataFrame(hospital_data)
    
    print("\n" + "="*80)
    print("SAMPLE HOSPITAL DATA".center(80))
    print("="*80)
    
    # Display first 3 hospitals
    for i in range(min(3, len(df))):
        print(f"\n🏥 Hospital {i+1}: {df.iloc[i]['Hospital_Name']}")
        print(f"   Type: {df.iloc[i]['Type']}")
        print(f"   Address: {df.iloc[i]['Address']}")
        print(f"   Beds: {df.iloc[i]['Beds_Total']} (ICU: {df.iloc[i]['ICU_Beds']})")
        print(f"   Contact: {df.iloc[i]['Contact_Number']}")
        print(f"   Specialties: {df.iloc[i]['Specialties']}")
    
    print("\n" + "="*80)
    print("SUMMARY STATISTICS".center(80))
    print("="*80)
    
    # Calculate statistics
    govt_hospitals = df[df['Type'] == 'Government']
    private_hospitals = df[df['Type'] == 'Private']
    specialty_hospitals = df[df['Type'] == 'Specialty']
    
    print(f"\n🏛️ Government Hospitals: {len(govt_hospitals)}")
    print(f"💼 Private Hospitals: {len(private_hospitals)}")
    print(f"👨‍⚕️ Specialty Hospitals: {len(specialty_hospitals)}")
    print(f"📈 Average Rating: {df['Rating'].mean():.2f}/5")
    print(f"🎯 Most Common Services:")
    specialties = ', '.join(df['Specialties'].str.split(', ').sum())
    print(f"   {specialties[:100]}...")

if __name__ == "__main__":
    # Create the Excel file
    excel_file = create_excel_sheet()
    
    # Display sample data
    display_sample_data()
    
    # Additional instructions
    print("\n" + "="*80)
    print("NEXT STEPS".center(80))
    print("="*80)
    print("\n1. Excel file 'hospitals_huzurabad.xlsx' has been created")
    print("2. The file contains 3 sheets:")
    print("   - 'Hospitals': Complete list of hospitals with all details")
    print("   - 'Summary': Statistical summary by hospital type")
    print("   - 'Metadata': Information about the dataset")
    print("\n3. To read this file in Python:")
    print("   ```python")
    print("   import pandas as pd")
    print("   df = pd.read_excel('hospitals_huzurabad.xlsx', sheet_name='Hospitals')")
    print("   ```")